# Description: CSV 파일을 엑셀 파일로 변환하고 첫 번째 행을 필터(테이블 헤더)로 설정

import pandas as pd
import argparse
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
import os

def convert_csv_to_excel(input_csv):
    """CSV 파일을 엑셀 파일로 변환하고 첫 번째 행을 필터(테이블 헤더)로 설정"""
    
    # 입력 파일 확인
    if not os.path.exists(input_csv):
        print(f"❌ 오류: 입력 파일 '{input_csv}'이(가) 존재하지 않습니다.")
        return

    # 출력 파일명 설정 (.xlsx로 변환)
    output_excel = os.path.splitext(input_csv)[0] + ".xlsx"

    # CSV 파일 읽기
    df = pd.read_csv(input_csv, encoding='utf-8')

    # 엑셀 파일로 저장 (첫 번째 행을 헤더로 사용)
    df.to_excel(output_excel, index=False, sheet_name="Sheet1")

    # 엑셀 파일 불러오기
    wb = load_workbook(output_excel)
    ws = wb.active

    # 데이터 범위 가져오기 (A1부터 마지막 열, 마지막 행까지)
    last_column = ws.max_column
    last_row = ws.max_row
    table_range = f"A1:{chr(64 + last_column)}{last_row}"

    # 테이블 생성 및 스타일 적용
    table = Table(displayName="DataTable", ref=table_range)
    style = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False
    )
    table.tableStyleInfo = style
    ws.add_table(table)

    # 엑셀 파일 저장
    wb.save(output_excel)

    print(f"✅ 엑셀 변환 완료! '{output_excel}'에 저장되었습니다.")

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="CSV 파일을 엑셀 파일로 변환하고 첫 번째 행을 필터(테이블 헤더)로 설정합니다.")
    parser.add_argument("input_csv", help="입력 CSV 파일 경로")
    
    args = parser.parse_args()
    
    convert_csv_to_excel(args.input_csv)
